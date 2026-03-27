"""
auditor.py — Core SEO crawling and analysis engine.
Extracted from master_analyser.py; no Flask/UI dependencies.
"""
from __future__ import annotations

import re
import time
import os
from collections import deque
from datetime import datetime
from typing import Callable, Optional
from urllib.parse import urlparse, urljoin, urlunparse

try:
    import textstat
    TEXTSTAT_AVAILABLE = True
except ImportError:
    TEXTSTAT_AVAILABLE = False

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

import io

# ---------------------------------------------------------------------------
# Browser user-agent
# ---------------------------------------------------------------------------
_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# ---------------------------------------------------------------------------
# WAF / challenge-page detection
# ---------------------------------------------------------------------------
_CHALLENGE_PHRASES = frozenset([
    "human verification", "verify you are human", "just a moment",
    "ddos protection", "checking your browser", "please wait",
    "access denied", "security check", "attention required",
    "one more step", "bot detection", "ray id", "enable javascript",
    "you have been blocked", "cf-browser-verification",
])

_SKELETON_SELECTORS = (
    '[class*="skeleton"]', '[class*="Skeleton"]', '[class*="shimmer"]',
    '[class*="loading"]', '[class*="placeholder"]', '[aria-busy="true"]',
    '.spinner', '[data-loading="true"]', '[class*="Spinner"]',
)


def _is_challenge_title(title: str) -> bool:
    t = title.lower()
    return any(ph in t for ph in _CHALLENGE_PHRASES)


# ---------------------------------------------------------------------------
# DOM extraction script (runs inside the browser page)
# ---------------------------------------------------------------------------
_AUDIT_DOM_SCRIPT = """() => {
  function q(sel) { return document.querySelector(sel); }
  function qa(sel) { return [...document.querySelectorAll(sel)]; }
  function getMeta(name) {
    const el = q('meta[name="' + name + '"]') || q('meta[property="' + name + '"]');
    return el ? (el.getAttribute('content') || '').trim() : '';
  }

  const title = document.title || '';
  const canonical = (q('link[rel="canonical"]') || {}).href || '';
  const metaDescription = getMeta('description');
  const metaRobots = getMeta('robots');
  const metaKeywords = getMeta('keywords');
  const metaViewport = getMeta('viewport');

  const h1s = qa('h1').map(el => el.innerText.trim()).filter(Boolean).slice(0, 5);
  const h2s = qa('h2').map(el => el.innerText.trim()).filter(Boolean).slice(0, 10);
  const h3s = qa('h3').map(el => el.innerText.trim()).filter(Boolean).slice(0, 10);

  const imgs = qa('img');
  const missingAlt = imgs.filter(el => el.getAttribute('alt') === null).length;
  const emptyAlt   = imgs.filter(el => el.getAttribute('alt') === '').length;
  const altTooLong = imgs.filter(el => (el.getAttribute('alt') || '').length > 125).length;

  const host = window.location.hostname;
  const allLinks = qa('a[href]').map(el => {
    const href = el.href || '';
    let isInternal = false;
    try { isInternal = (new URL(href)).hostname === host; } catch(e) { isInternal = !href.startsWith('http'); }
    return {
      href: href.substring(0, 500),
      text: (el.innerText || '').trim().substring(0, 200),
      rel: el.getAttribute('rel') || '',
      isInternal,
      isNofollow: (el.getAttribute('rel') || '').includes('nofollow'),
    };
  });
  const skip = h => h.startsWith('javascript') || h.startsWith('mailto') || h.startsWith('tel') || h === '';
  const internalLinks = allLinks.filter(l => l.isInternal && !skip(l.href));
  const externalLinks = allLinks.filter(l => !l.isInternal && !skip(l.href));
  const nofollowLinks = allLinks.filter(l => l.isNofollow);

  const og = {};
  ['title','description','image','type','url','site_name'].forEach(p => { og[p] = getMeta('og:' + p); });
  const twitter = {};
  ['card','title','description','image','site'].forEach(p => { twitter[p] = getMeta('twitter:' + p); });

  const schemas = qa('script[type="application/ld+json"]').map(el => {
    try {
      const d = JSON.parse(el.textContent || '');
      return Array.isArray(d) ? d.map(x => x['@type']) : [d['@type']];
    } catch(e) { return []; }
  }).flat().filter(Boolean);

  const hreflang = qa('link[hreflang]').map(el => el.getAttribute('hreflang')).filter(Boolean);
  const relNext = ((q('link[rel="next"]') || {}).href || '');
  const relPrev = ((q('link[rel="prev"]') || {}).href || '');

  const cloneBody = document.body ? document.body.cloneNode(true) : null;
  if (cloneBody) {
    cloneBody.querySelectorAll('script,style,nav,header,footer,aside').forEach(el => el.remove());
  }
  const bodyText = ((cloneBody || document.body || {}).innerText || '').trim();
  const wordCount = bodyText.split(/\\s+/).filter(w => w.length > 0).length;

  const numScripts     = qa('script[src]').length;
  const numStylesheets = qa('link[rel="stylesheet"]').length;
  const numIframes     = qa('iframe').length;
  const paragraphCount = qa('p').length;
  const htmlSize       = (document.documentElement || {}).outerHTML
                          ? document.documentElement.outerHTML.length : 0;
  const textToHtmlRatio = htmlSize > 0
    ? parseFloat(((bodyText.length / htmlSize) * 100).toFixed(2)) : 0;

  const hasBreadcrumbs = (
    qa('[class*="breadcrumb"]').length > 0 ||
    qa('nav[aria-label*="breadcrumb" i]').length > 0 ||
    schemas.includes('BreadcrumbList')
  );

  return {
    url: window.location.href, title, canonical, metaDescription, metaRobots,
    metaKeywords, metaViewport,
    h1s, h2s, h3s,
    imageCount: imgs.length, missingAlt, emptyAlt, altTooLong,
    internalLinkCount: internalLinks.length,
    externalLinkCount: externalLinks.length,
    nofollowLinkCount: nofollowLinks.length,
    totalLinkCount: allLinks.length,
    internalLinks: internalLinks.slice(0, 400).map(l => l.href),
    externalLinksData: externalLinks.slice(0, 50).map(l => ({href: l.href, text: l.text, nofollow: l.isNofollow})),
    og, twitter,
    schemaTypes: schemas,
    hreflang, relNext, relPrev,
    wordCount, paragraphCount, numScripts, numStylesheets, numIframes,
    htmlSize, textToHtmlRatio, hasBreadcrumbs,
    bodyText: bodyText.substring(0, 5000),
  };
}"""


# ---------------------------------------------------------------------------
# Content-ready waiter
# ---------------------------------------------------------------------------
def _wait_for_real_content(page, min_chars: int = 800, timeout_s: int = 30) -> dict:
    """
    After network-idle, actively wait for JS-rendered content to fully appear.
    Handles React/Next.js hydration, lazy-loaded components, and WAF challenges.
    """
    try:
        for frac in [0.25, 0.5, 0.75, 1.0]:
            page.evaluate(
                f"window.scrollTo(0, document.documentElement.scrollHeight * {frac})"
            )
            page.wait_for_timeout(700)
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(500)
    except Exception:
        pass

    skeleton_js = (
        "() => {"
        "  const sels = " + str(list(_SKELETON_SELECTORS)) + ";"
        "  return sels.some(s => {"
        "    try {"
        "      const el = document.querySelector(s);"
        "      return el && el.offsetParent !== null;"
        "    } catch(e) { return false; }"
        "  });"
        "}"
    )
    text_js = (
        "() => {"
        "  const b = document.body; if (!b) return 0;"
        "  const c = b.cloneNode(true);"
        "  c.querySelectorAll('script,style,noscript,template').forEach(e=>e.remove());"
        "  return (c.innerText || c.textContent || '').trim().length;"
        "}"
    )

    deadline = time.time() + timeout_s
    last_len = -1
    stall_ticks = 0

    while time.time() < deadline:
        try:
            title = page.evaluate("() => document.title || ''")
            char_count = page.evaluate(text_js)
            has_skeleton = page.evaluate(skeleton_js)
        except Exception:
            break

        if _is_challenge_title(title):
            page.wait_for_timeout(5000)
            try:
                new_title = page.evaluate("() => document.title || ''")
            except Exception:
                new_title = title
            if _is_challenge_title(new_title):
                return {"is_challenge": True, "timed_out": False, "final_chars": char_count}
            try:
                for frac in [0.5, 1.0, 0.0]:
                    page.evaluate(
                        f"window.scrollTo(0, document.documentElement.scrollHeight * {frac})"
                    )
                    page.wait_for_timeout(700)
            except Exception:
                pass
            continue

        if char_count >= min_chars and not has_skeleton:
            return {"is_challenge": False, "timed_out": False, "final_chars": char_count}

        if char_count == last_len:
            stall_ticks += 1
            if stall_ticks >= 5:
                break
        else:
            stall_ticks = 0
        last_len = char_count
        page.wait_for_timeout(1200)

    try:
        title = page.evaluate("() => document.title || ''")
        char_count = page.evaluate(text_js)
    except Exception:
        title, char_count = "", 0
    return {"is_challenge": _is_challenge_title(title), "timed_out": True, "final_chars": char_count}


# ---------------------------------------------------------------------------
# Build result dict from raw data
# ---------------------------------------------------------------------------
def _build_audit_result(url, final_url, status_code, http_headers,
                        response_time_ms, html_size_bytes, dom_data,
                        ttfb_ms: int = 0, full_load_ms: int = 0):
    h1s = dom_data.get("h1s") or []
    h2s = dom_data.get("h2s") or []
    h3s = dom_data.get("h3s") or []
    title = dom_data.get("title") or ""
    meta_desc = dom_data.get("metaDescription") or ""
    schemas = dom_data.get("schemaTypes") or []
    og = dom_data.get("og") or {}
    twitter = dom_data.get("twitter") or {}
    hreflang = dom_data.get("hreflang") or []

    result = {
        "URL": url,
        "Final URL": final_url,
        "Status Code": status_code,
        "Redirect URL": final_url if final_url != url else "",
        # TTFB = HTTP response received (what Screaming Frog calls "Response Time")
        # Full Load = TTFB + JS hydration wait (time until real content appeared)
        "Response Time (ms)": ttfb_ms or response_time_ms,
        "Full Load Time (ms)": full_load_ms or response_time_ms,
        "Content Type": http_headers.get("content-type", ""),
        "Page Size (bytes)": html_size_bytes,
        "Page Size (KB)": round(html_size_bytes / 1024, 1),
        "X-Robots-Tag": http_headers.get("x-robots-tag", ""),
        "Last Modified": http_headers.get("last-modified", ""),
        "Server": http_headers.get("server", ""),
        "HTTPS": url.startswith("https"),
        "Title": title,
        "Title Length": len(title),
        "Meta Description": meta_desc,
        "Meta Description Length": len(meta_desc),
        "Meta Keywords": dom_data.get("metaKeywords") or "",
        "Meta Robots": dom_data.get("metaRobots") or "",
        "Meta Viewport": dom_data.get("metaViewport") or "",
        "H1 Count": len(h1s),
        "H1 First": h1s[0] if h1s else "",
        "H2 Count": len(h2s),
        "H2 First": h2s[0] if h2s else "",
        "H3 Count": len(h3s),
        "Canonical URL": dom_data.get("canonical") or "",
        "Image Count": dom_data.get("imageCount") or 0,
        "Images Missing Alt": dom_data.get("missingAlt") or 0,
        "Images Empty Alt": dom_data.get("emptyAlt") or 0,
        "Internal Links": dom_data.get("internalLinkCount") or 0,
        "External Links": dom_data.get("externalLinkCount") or 0,
        "Nofollow Links": dom_data.get("nofollowLinkCount") or 0,
        "Total Links": dom_data.get("totalLinkCount") or 0,
        "OG Title": og.get("title") or "",
        "OG Description": og.get("description") or "",
        "OG Image": og.get("image") or "",
        "OG Type": og.get("type") or "",
        "Twitter Card": twitter.get("card") or "",
        "Schema Types": ", ".join(schemas),
        "Has Structured Data": bool(schemas),
        "Hreflang Languages": ", ".join(hreflang),
        "Has Hreflang": bool(hreflang),
        "Word Count": dom_data.get("wordCount") or 0,
        "Paragraph Count": dom_data.get("paragraphCount") or 0,
        "HTML Size (bytes)": dom_data.get("htmlSize") or html_size_bytes,
        "Text to HTML Ratio (%)": dom_data.get("textToHtmlRatio") or 0,
        "Scripts Count": dom_data.get("numScripts") or 0,
        "Stylesheets Count": dom_data.get("numStylesheets") or 0,
        "Iframes Count": dom_data.get("numIframes") or 0,
        "Flesch Reading Ease": "",
        "Indexable": True,
        "Indexability Issues": "",
        "WAF Blocked": dom_data.get("_waf_challenge", False),
        "Content Timed Out": dom_data.get("_content_timed_out", False),
        "_internal_links": dom_data.get("internalLinks") or [],
        "_body_text": dom_data.get("bodyText") or "",
    }

    body_text = dom_data.get("bodyText") or ""
    if body_text and len(body_text.split()) > 30 and TEXTSTAT_AVAILABLE:
        try:
            result["Flesch Reading Ease"] = round(textstat.flesch_reading_ease(body_text), 1)
        except Exception:
            pass

    meta_robots_lower = (result["Meta Robots"] + " " + result["X-Robots-Tag"]).lower()
    noindex = "noindex" in meta_robots_lower
    result["Indexable"] = not noindex and status_code in (200, 0)
    if noindex:
        result["Indexability Issues"] = "noindex directive"
    elif status_code not in (200, 0):
        result["Indexability Issues"] = f"HTTP {status_code}"

    issues = _classify_seo_issues(result)
    result["Critical Issues"] = "; ".join(issues["critical"])
    result["Warnings"] = "; ".join(issues["warnings"])
    result["Info"] = "; ".join(issues["info"])
    result["Critical Count"] = len(issues["critical"])
    result["Warning Count"] = len(issues["warnings"])
    result["Info Count"] = len(issues["info"])

    return result


# ---------------------------------------------------------------------------
# Issue classifier
# ---------------------------------------------------------------------------
def _classify_seo_issues(data: dict) -> dict:
    critical, warnings, info = [], [], []

    waf_blocked = data.get("WAF Blocked", False)
    content_timed_out = data.get("Content Timed Out", False)

    if waf_blocked:
        critical.append("WAF / bot-challenge page — real content was not accessible")
        if not data.get("HTTPS"):
            critical.append("Page served over HTTP (not HTTPS)")
        sc = data.get("Status Code") or 0
        if sc == 404:
            critical.append("Page returns 404 (not found)")
        elif sc >= 500:
            critical.append(f"Server error ({sc})")
        rt = data.get("Response Time (ms)") or 0
        if rt > 3000:
            warnings.append(f"Slow response time ({rt} ms)")
        return {"critical": critical, "warnings": warnings, "info": info}

    if content_timed_out:
        warnings.append("Page content may be incomplete — JS rendering timed out")

    title = data.get("Title") or ""
    tlen = len(title)
    if not title:
        critical.append("Missing title tag")
    elif tlen > 60:
        warnings.append(f"Title too long ({tlen} chars, recommended ≤60)")
    elif tlen < 30:
        warnings.append(f"Title too short ({tlen} chars, recommended ≥30)")

    desc = data.get("Meta Description") or ""
    dlen = len(desc)
    if not desc:
        warnings.append("Missing meta description")
    elif dlen > 160:
        warnings.append(f"Meta description too long ({dlen} chars, recommended ≤160)")
    elif dlen < 70:
        info.append(f"Meta description short ({dlen} chars, recommended ≥70)")

    h1 = data.get("H1 Count") or 0
    if h1 == 0:
        critical.append("Missing H1 tag")
    elif h1 > 1:
        warnings.append(f"Multiple H1 tags ({h1})")

    missing_alt = data.get("Images Missing Alt") or 0
    if missing_alt > 0:
        warnings.append(f"{missing_alt} image(s) missing alt text")

    canonical = data.get("Canonical URL") or ""
    url = data.get("URL") or ""
    final_url = data.get("Final URL") or url
    if not canonical:
        warnings.append("Missing canonical tag")
    elif canonical not in (url, final_url) and not final_url.startswith(canonical.rstrip("/")):
        info.append(f"Canonical points elsewhere ({canonical[:60]})")

    meta_robots = (data.get("Meta Robots") or "").lower()
    x_robots = (data.get("X-Robots-Tag") or "").lower()
    if "noindex" in meta_robots or "noindex" in x_robots:
        critical.append("noindex directive — page will not be indexed")
    if "nofollow" in meta_robots:
        warnings.append("Meta robots nofollow — link equity not passed")

    if not data.get("HTTPS"):
        critical.append("Page served over HTTP (not HTTPS)")

    sc = data.get("Status Code") or 0
    if sc in (301, 302, 307, 308):
        info.append(f"Redirect ({sc}) → {data.get('Final URL', '')[:80]}")
    elif sc == 404:
        critical.append("Page returns 404 (not found)")
    elif sc >= 500:
        critical.append(f"Server error ({sc})")
    elif sc == 403:
        warnings.append("403 Forbidden")
    elif sc not in (0, 200, 201) and sc not in (301, 302, 307, 308):
        warnings.append(f"Unexpected HTTP status {sc}")

    if not (data.get("OG Title") or ""):
        info.append("Missing og:title")
    if not (data.get("OG Description") or ""):
        info.append("Missing og:description")
    if not (data.get("OG Image") or ""):
        info.append("Missing og:image")

    if not data.get("Has Structured Data"):
        info.append("No Schema.org structured data found")

    page_kb = (data.get("Page Size (bytes)") or 0) / 1024
    if page_kb > 5000:
        warnings.append(f"Very large page ({page_kb:.0f} KB)")
    elif page_kb > 2000:
        info.append(f"Page size {page_kb:.0f} KB (consider optimisation)")

    rt = data.get("Response Time (ms)") or 0          # TTFB
    fl = data.get("Full Load Time (ms)") or rt         # full JS load
    if rt > 3000:
        warnings.append(f"Slow server response / TTFB ({rt} ms)")
    elif rt > 1500:
        info.append(f"Server response time {rt} ms (TTFB)")
    if fl > 5000 and fl != rt:
        info.append(f"Full page load took {fl} ms (JS rendering included)")

    wc = data.get("Word Count") or 0
    if 0 < wc < 100:
        warnings.append(f"Thin content ({wc} words)")

    if not data.get("Has Hreflang"):
        info.append("No hreflang tags (may be intentional for single-language sites)")

    return {"critical": critical, "warnings": warnings, "info": info}


# ---------------------------------------------------------------------------
# Browser context factory
# ---------------------------------------------------------------------------
def _new_stealth_context(browser):
    ctx = browser.new_context(
        user_agent=_UA,
        viewport={"width": 1440, "height": 900},
        locale="en-US",
        extra_http_headers={
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "DNT": "1",
        },
    )
    return ctx


# ---------------------------------------------------------------------------
# Single-page audit
# ---------------------------------------------------------------------------
def audit_page(url: str) -> dict:
    """Audit a single URL using a stealth headless browser."""
    if not PLAYWRIGHT_AVAILABLE:
        return {
            "URL": url, "Status Code": 0,
            "Critical Issues": "Playwright not installed",
            "Warnings": "", "Info": "",
            "Critical Count": 1, "Warning Count": 0, "Info Count": 0,
        }
    start = time.time()
    try:
        with sync_playwright() as p:
            try:
                browser = p.chromium.launch(
                    channel="chrome", headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox",
                          "--disable-dev-shm-usage"],
                )
            except Exception:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox",
                          "--disable-dev-shm-usage"],
                )
            ctx = _new_stealth_context(browser)
            page = ctx.new_page()
            page.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
            )

            status_code = [200]
            http_headers = [{}]

            def _on_response(resp):
                if resp.url == url or resp.url == page.url:
                    status_code[0] = resp.status
                    http_headers[0] = dict(resp.headers)

            page.on("response", _on_response)
            t0 = time.time()
            resp = page.goto(url, wait_until="domcontentloaded", timeout=30000)
            if resp:
                status_code[0] = resp.status
                http_headers[0] = dict(resp.headers)

            # TTFB via Navigation Timing API (most accurate)
            ttfb = 0
            try:
                ttfb = page.evaluate(
                    "() => { try { const [n] = performance.getEntriesByType('navigation');"
                    " return n ? Math.round(n.responseStart) : 0; } catch(e) { return 0; } }"
                )
            except Exception:
                pass

            # Fallback: DOMContentLoaded time if TTFB not available
            domcl_time = round((time.time() - t0) * 1000)
            ttfb = ttfb or domcl_time

            try:
                page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass

            content_check = _wait_for_real_content(page, min_chars=600, timeout_s=30)
            full_load_time = round((time.time() - t0) * 1000)
            final_url = page.url
            dom_data = page.evaluate(_AUDIT_DOM_SCRIPT)
            dom_data["_waf_challenge"] = content_check["is_challenge"]
            dom_data["_content_timed_out"] = content_check["timed_out"]
            html_content = page.content()
            html_size = len(html_content.encode("utf-8"))
            browser.close()

        return _build_audit_result(url, final_url, status_code[0], http_headers[0],
                                   ttfb, html_size, dom_data,
                                   ttfb_ms=ttfb, full_load_ms=full_load_time)
    except Exception as e:
        return {
            "URL": url, "Final URL": url, "Status Code": 0,
            "Critical Issues": f"Crawl failed: {str(e)[:200]}",
            "Warnings": "", "Info": "",
            "Critical Count": 1, "Warning Count": 0, "Info Count": 0,
        }


# ---------------------------------------------------------------------------
# Site crawler (BFS)
# ---------------------------------------------------------------------------
def crawl_site(
    start_url: str,
    max_pages: int = 50,
    delay_s: float = 1.0,
    progress_callback: Optional[Callable] = None,
) -> list[dict]:
    """BFS-crawl a site; audit each internal page up to max_pages."""
    if not PLAYWRIGHT_AVAILABLE:
        return []

    parsed_start = urlparse(start_url)
    base_domain = parsed_start.netloc
    queue: deque = deque([start_url])
    visited: set = set()
    results: list = []

    try:
        with sync_playwright() as p:
            try:
                browser = p.chromium.launch(
                    channel="chrome", headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox",
                          "--disable-dev-shm-usage"],
                )
            except Exception:
                browser = p.chromium.launch(
                    headless=True,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox",
                          "--disable-dev-shm-usage"],
                )
            ctx = _new_stealth_context(browser)

            while queue and len(visited) < max_pages:
                url = queue.popleft()
                if url in visited:
                    continue
                visited.add(url)

                if progress_callback:
                    progress_callback(len(visited), max_pages, url)

                page = ctx.new_page()
                page.add_init_script(
                    "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                )

                status_code = [200]
                http_headers = [{}]

                def _on_resp(resp, _url=url):
                    if resp.url == _url:
                        status_code[0] = resp.status
                        http_headers[0] = dict(resp.headers)

                page.on("response", _on_resp)

                try:
                    t0 = time.time()
                    resp = page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    if resp:
                        status_code[0] = resp.status
                        http_headers[0] = dict(resp.headers)

                    ttfb = 0
                    try:
                        ttfb = page.evaluate(
                            "() => { try { const [n] = performance.getEntriesByType('navigation');"
                            " return n ? Math.round(n.responseStart) : 0; } catch(e) { return 0; } }"
                        )
                    except Exception:
                        pass
                    domcl_time = round((time.time() - t0) * 1000)
                    ttfb = ttfb or domcl_time

                    try:
                        page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass

                    content_check = _wait_for_real_content(page, min_chars=600, timeout_s=25)
                    full_load_time = round((time.time() - t0) * 1000)
                    final_url = page.url
                    dom_data = page.evaluate(_AUDIT_DOM_SCRIPT)
                    dom_data["_waf_challenge"] = content_check["is_challenge"]
                    dom_data["_content_timed_out"] = content_check["timed_out"]
                    html_content = page.content()
                    html_size = len(html_content.encode("utf-8"))

                    data = _build_audit_result(url, final_url, status_code[0],
                                               http_headers[0], ttfb,
                                               html_size, dom_data,
                                               ttfb_ms=ttfb, full_load_ms=full_load_time)
                    data["Crawl Depth"] = url.count("/") - start_url.count("/")
                    results.append(data)

                    # Only follow links from non-WAF pages
                    if not data.get("WAF Blocked"):
                        for link in dom_data.get("internalLinks") or []:
                            pl = urlparse(link)
                            if pl.netloc == base_domain or not pl.netloc:
                                clean = urlunparse((
                                    pl.scheme or parsed_start.scheme,
                                    pl.netloc or base_domain,
                                    pl.path, pl.params, pl.query, ""
                                ))
                                if clean not in visited and clean not in queue:
                                    queue.append(clean)
                except Exception as e:
                    results.append({
                        "URL": url, "Status Code": 0,
                        "Critical Issues": f"Crawl error: {str(e)[:120]}",
                        "Warnings": "", "Info": "",
                        "Critical Count": 1, "Warning Count": 0, "Info Count": 0,
                    })
                finally:
                    page.close()

                if delay_s > 0:
                    time.sleep(delay_s)

            browser.close()
    except Exception as e:
        results.append({
            "URL": start_url, "Status Code": 0,
            "Critical Issues": f"Browser launch failed: {str(e)[:200]}",
            "Warnings": "", "Info": "",
            "Critical Count": 1, "Warning Count": 0, "Info Count": 0,
        })

    return results


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------
def export_excel(results: list[dict]) -> bytes:
    """Return an Excel file as bytes."""
    if not OPENPYXL_AVAILABLE or not results:
        return b""

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    total = len(results)
    critical_pages = sum(1 for r in results if (r.get("Critical Count") or 0) > 0)
    warning_pages = sum(1 for r in results if (r.get("Warning Count") or 0) > 0)
    indexable = sum(1 for r in results if r.get("Indexable"))
    waf_blocked = sum(1 for r in results if r.get("WAF Blocked"))
    avg_rt = round(sum(r.get("Response Time (ms)") or 0 for r in results) / max(total, 1))

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 20
    hdr_fill = PatternFill("solid", fgColor="1A3C5E")
    hdr_font = Font(color="FFFFFF", bold=True)
    for row in [
        ["Metric", "Value"],
        ["Total Pages Audited", total],
        ["Pages with Critical Issues", critical_pages],
        ["Pages with Warnings", warning_pages],
        ["Indexable Pages", indexable],
        ["Non-Indexable Pages", total - indexable],
        ["WAF-Blocked Pages", waf_blocked],
        ["Average Response Time (ms)", avg_rt],
    ]:
        ws_sum.append(row)
    for cell in ws_sum[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    # --- All Pages sheet ---
    ws = wb.create_sheet("All Pages")
    columns = [
        "URL", "Status Code", "Response Time (ms)", "Title", "Title Length",
        "Meta Description Length", "H1 Count", "Word Count", "Canonical URL",
        "Indexable", "WAF Blocked", "HTTPS", "Has Structured Data",
        "Image Count", "Images Missing Alt", "Internal Links",
        "Critical Count", "Warning Count", "Critical Issues", "Warnings", "Info",
    ]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    crit_fill = PatternFill("solid", fgColor="FFE0E0")
    warn_fill = PatternFill("solid", fgColor="FFF8E0")
    ok_fill = PatternFill("solid", fgColor="E8F8EE")
    waf_fill = PatternFill("solid", fgColor="F3E8FF")

    for r in results:
        row_data = [r.get(c, "") for c in columns]
        ws.append(row_data)
        row_idx = ws.max_row
        fill = (waf_fill if r.get("WAF Blocked") else
                crit_fill if (r.get("Critical Count") or 0) > 0 else
                warn_fill if (r.get("Warning Count") or 0) > 0 else ok_fill)
        for cell in ws[row_idx]:
            cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
